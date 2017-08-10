#!/usr/bin/env python
#coding:utf-8
import openpyxl,os,sys,time
x_file = sys.argv[1]
m_file = sys.argv[2]
new_time = time.strftime("%Y%m%d-%H:%M", time.localtime())
new_file = "%s.xlsx" % (new_time)
####分析所有发送数据发送数据
def Mos(name,phone='A',text='B',status='E'):
    #name:mos输出excel文件
    mos = openpyxl.load_workbook(name).worksheets[0]	#打开该文件的第一页
    mos_num = mos.max_row	#获取行数
    #a:号码,B:内容,E:发送状态
    L1,L2,L3 = "%s%s" % (phone,mos_num),"%s%s" % (text,mos_num),"%s%s" % (status,mos_num)
    M1,M2,M3 = "%s2" % (phone),"%s2" % (text),"%s2" % (status)
    ma,mb,mc = mos[M1:L1],mos[M2:L2],mos[M3:L3]
    mos_list = {}
    num = 0
    for x in range(0,mos_num-1):
        phone = ma[x][0].value
        text = mb[x][0].value
        status = mc[x][0].value
        value = {
		"phone" : phone,
		"text" : text,
		"status" : status
	}
        mos_list[num] = value
        num += 1
    return mos_list
def Xsd(name,phone_list,phone='M',text='E',user='D',num=1):
    xsd = openpyxl.load_workbook(name).worksheets[0]
    xsd_num = xsd.max_row
    X1,X2,X3 = "%s2" % (phone),"%s2" % (text),"%s2" % (user)
    N1,N2,N3 = "%s%s" % (phone,xsd_num),"%s%s" % (text,xsd_num),"%s%s" % (user,xsd_num)
    xsd_a,xsd_b,xsd_c = xsd[X1:N1],xsd[X2:N2],xsd[X3:N3]
    new_xls = excel.active
    new_xls['A1'] = "号码"
    new_xls['B1'] = "姓名"
    new_xls['C1'] = "发送状态"
    new_xls['D1'] = "发送内容"
    for y in range(0,xsd_num-1):
        user_list =  xsd_c[y][0].value
        phone = xsd_a[y][0].value
        text = xsd_b[y][0].value
        if phone != None:
            for y in range(0,len(phone_list)):
                if phone == phone_list[y]['phone'] and text in phone_list[y]['text'] and phone_list[y]['status'] != u"发送成功":
                    num += 1
                    A,B,C,D = "A%s" % (num),"B%s" % (num),"C%s" % (num),"D%s" % (num)
                    new_xls[A] = phone
                    new_xls[B] = user_list
                    new_xls[C] = phone_list[y]['status']
                    new_xls[D] = text
    return num
def status_num(name,num):
    excela = excel.active
    aa = bb = cc = 0
    for x in range(0,num):
        qq = "C%s" % (num)
        ma = excela['C1':qq]
        if ma[x][0].value == u"发送失败":
            aa += 1
        elif ma[x][0].value == u"已提交":
            bb += 1
    excela["A%s"%(num+2)] = "发送失败"
    excela["B%s"%(num+2)] = aa
    excela["C%s"%(num+2)] = "提交成功"
    excela["D%s"%(num+2)] = bb
    excel.save(new_file)
    return None
if __name__ == '__main__':
    L = Mos(m_file)
    excel = openpyxl.Workbook()
    num = Xsd(x_file,L)
    excel.save(new_file)
    status_num(new_file,num)
